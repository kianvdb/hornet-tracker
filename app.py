#!/usr/bin/env python3
"""
Hornet Tracker - Ground Station Server (v3 "ultimate")

Strategie:
  - Persistent RTL-SDR (geen rtl_power subprocess, geen re-init)
  - Peak-hold detectie over 400ms window -> pakt beacon packets (~200ms TX)
  - Vaste 1-malige baseline (uit v1, werkte het beste)
  - Smalle 200 kHz detectie-band -> minder ruis buiten
  - Abstracte SignalSource interface -> Ra-01 LoRa ontvanger plug-in klaar

Toekomstige LoRa ontvanger (Ra-01 via SPI):
  1. Implementeer LoRaSource class (unten in file, nu als stub)
  2. Zet SIGNAL_SOURCE = 'lora' bovenaan
  3. Restart service -> dashboard werkt identiek
  Alles daarbuiten (WiFi, MAVLink, dashboard) blijft onveranderd.

Kalibratie van beacon v2 (zie beacon.ino):
  - SF9, BW 125 kHz, 433 MHz, ~200ms time-on-air, elke 500ms
"""
import mission  # onze demo-missie module
from flask import Flask, render_template, jsonify, request, send_file
import io
import json
import math
from flask_socketio import SocketIO
import subprocess
import threading
import time
import os
import re
import numpy as np

# ============================================
# CONFIGURATIE
# ============================================

# --- Bron selectie (FUTURE-PROOF) ---
# 'rtlsdr' = huidige RTL-SDR detectie (energy detection)
# 'lora'   = Ra-01 SPI ontvanger (packet decoding) -- nog te implementeren
SIGNAL_SOURCE = 'lora'

# --- RTL-SDR parameters ---
CENTER_FREQ = 433.0e6        # Beacon frequentie (Hz)
SAMPLE_RATE = 2.048e6        # RTL-SDR native rate (Hz)
GAIN = 49.6                  # dB
NUM_SAMPLES = 16384          # per meting (~8ms data)

DETECT_BANDWIDTH = 200e3     # Smal band rond center -> filter ruis
THRESHOLD = 6                # dB boven baseline = detectie

# --- Peak-hold parameters ---
# Beacon v2 zendt ~200ms elke 500ms, dus window van 400ms pakt hem gegarandeerd
PEAK_HOLD_SAMPLES = 50       # 50 x 8ms = ~400ms window
MEASURE_INTERVAL = 0.1       # 10 Hz update naar dashboard

# --- Baseline (vast, na initiele meting) ---
BASELINE_NUM_READINGS = 10   # aantal metingen voor initiele baseline

# --- MAVLink ---
# Directe USB-verbinding naar de Pixhawk (interface 0 = MAVLink telemetrie).
# We gebruiken bewust NIET de USB-TTL naar TELEM2 (/dev/ttyUSB0, CH340) of de
# SiK-radio: die zijn trager en onbetrouwbaar gebleken. De directe USB-poort
# is CDC-ACM, dus de baudrate wordt genegeerd — 115200 is een formaliteit.
MAVLINK_DEVICE = '/dev/serial/by-id/usb-Holybro_Pixhawk6C_32003E001651333337363133-if00'
MAVLINK_BAUD = 115200

# --- WiFi hotspot ---
WIFI_HOTSPOT_IFACE = 'wlan1'

# --- Thermal camera (Pimoroni MLX90640 op I2C) ---
# Refresh rate: 16 Hz = effectief ~8 FPS na I2C-overhead.
# Bij issues (skipped frames, "Too many retries"): verlaag naar 8 of 4.
THERMAL_REFRESH_HZ = 8
THERMAL_EMIT_INTERVAL = 0.15  # max ~6-7 emits/sec naar dashboard

# --- Persistente log storage ---
# JSON-bestand op disk i.p.v. browser-localStorage zodat log overleeft
# browser-refresh, andere browsers, andere devices, en Pi-reboots.
# Atomic writes via temp-file + os.replace zodat een crash midden in
# een write geen corrupte file achterlaat.
DATA_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data')
LOG_FILE = os.path.join(DATA_DIR, 'coord-log.json')
log_lock = threading.Lock()

# --- Tile cache ---
# Map waar offline tiles worden opgeslagen, gestructureerd als
#   data/tiles/<source>/<z>/<x>/<y>.png
# Twee sources: 'osm' (OpenStreetMap stratenplan) en 'sat' (ArcGIS satelliet)
TILE_CACHE_DIR = os.path.join(DATA_DIR, 'tiles')

# Externe tile-servers — gebruikt als fallback bij cache-miss met internet.
# {z}/{x}/{y} placeholders worden vervangen door python format string.
TILE_SOURCES = {
    'osm': 'https://tile.openstreetmap.org/{z}/{x}/{y}.png',
    'sat': 'https://server.arcgisonline.com/ArcGIS/rest/services/World_Imagery/MapServer/tile/{z}/{y}/{x}',
    'hyb': 'https://server.arcgisonline.com/ArcGIS/rest/services/Reference/World_Transportation/MapServer/tile/{z}/{y}/{x}',
    # NB: ArcGIS gebruikt {z}/{y}/{x} (omgekeerde volgorde), niet typo
}

# User-agent vereist door OSM tile-server policy. Beleefd identificeren
# voorkomt dat ze ons IP blokkeren bij prefetch van veel tiles.
TILE_USER_AGENT = 'VespaTrack/1.0 (bachelorthesis Erasmushogeschool Brussel)'

# ============================================
# FLASK APP
# ============================================
app = Flask(__name__)
app.config['SECRET_KEY'] = 'hornet-tracker-secret'
socketio = SocketIO(app, cors_allowed_origins="*", async_mode='threading')

status = {
    'signal_power': -100,
    'signal_delta': 0,
    'signal_detected': False,
    'baseline': -100,
    'signal_source': SIGNAL_SOURCE,  # dashboard kan tonen welke bron actief is

    'pi_connected': True,
    'pixhawk_connected': False,

    'gps_lat': 0, 'gps_lon': 0, 'gps_fix': False, 'gps_satellites': 0,
    'altitude': 0,
    'heading': 0,
    'battery_voltage': 0, 'battery_percent': 0,
    'flight_mode': 'UNKNOWN', 'armed': False,

    'wifi_connected': False, 'wifi_rssi': -100, 'wifi_quality': 0, 'wifi_clients': 0,

    'telem_connected': False,
    'telem_rssi_local': 0, 'telem_rssi_remote': 0,
    'telem_noise_local': 0, 'telem_noise_remote': 0,
    'telem_quality': 0, 'telem_txbuf': 0,

# LoRa-specifieke velden (stub voor toekomst)
    'lora_packet_count': 0,
    'lora_last_tracker_id': 0,
    'lora_last_seen_sec': -1,
    'lora_snr': 0,

    # Thermal camera status (frame-data wordt apart via 'thermal_frame' emit gestuurd
    # om de status_update payload klein te houden)
'thermal_connected': False,
    'thermal_min': 0.0,
    'thermal_max': 0.0,
    'thermal_avg': 0.0,
    'thermal_fps': 0.0,
    # Baseline detectie-modus: operator drukt op "baseline instellen" knop
    # in dashboard om het huidige frame als referentie op te slaan. Daarna
    # kan de frontend kiezen om alleen pixels te tonen die boven baseline liggen.
    'thermal_baseline_set': False,
}


# Baseline frame buiten status dict — 768 floats hoort niet in status_update payload.
# Wordt meegestuurd in thermal_frame events zodra ingesteld.
thermal_baseline = None  # None of list[768] floats
_last_thermal_frame = None  # laatst gelezen frame, voor baseline-capture
baseline_reset_requested = False

# Globale MAVLink connectie - gezet door mavlink_loop() na succesvolle connect.
# Wordt gebruikt door command handlers (arm/disarm/set_mode) om commando's
# naar de Pixhawk te sturen. None betekent: geen verbinding actief.
mav_connection = None
mav_lock = threading.Lock()

def get_mav_connection():
    """
    Getter voor de gedeelde MAVLink-connectie, gebruikt door mission.py.
    Returnt de huidige connectie of None. Thread-safe via mav_lock.
    """
    with mav_lock:
        return mav_connection

class AckMailbox:
    """
    Thread-safe 'postbus' voor COMMAND_ACK-berichten.
 
    Werking:
      - Een handler die een ack verwacht roept arm(command_id) aan vlak
        VOORDAT hij het commando verstuurt. Dat zet de verwachting klaar.
      - mavlink_loop() roept deliver(command_id, result) aan zodra er een
        COMMAND_ACK binnenkomt. Alleen als het command_id matcht met wat de
        handler verwacht, wordt de ack afgeleverd en de handler gewekt.
      - De handler roept wait(timeout) aan: blokkeert tot de ack er is of
        de timeout verstrijkt.
 
    We ondersteunen bewust maar één wachtende ack tegelijk. Dat is genoeg:
    de dashboard-knoppen worden nooit twee tegelijk ingedrukt, en de missie
    stuurt sequentieel. Zo blijft het simpel en voorspelbaar.
    """
    def __init__(self):
        self._lock = threading.Lock()
        self._event = threading.Event()
        self._expected_command = None   # command_id waarop we wachten
        self._result = None             # MAV_RESULT code van de ack
 
    def arm(self, command_id):
        """Zet klaar dat we een ack voor dit command_id verwachten."""
        with self._lock:
            self._expected_command = command_id
            self._result = None
            self._event.clear()
 
    def deliver(self, command_id, result):
        """
        Aangeroepen door mavlink_loop bij een COMMAND_ACK. Levert af als het
        command_id matcht met de verwachting. Returnt True als afgeleverd.
        """
        with self._lock:
            if self._expected_command is not None and command_id == self._expected_command:
                self._result = result
                self._event.set()
                return True
        return False
 
    def wait(self, timeout):
        """
        Wacht tot de verwachte ack binnen is (of timeout). Returnt de
        MAV_RESULT code, of None bij timeout.
        """
        got = self._event.wait(timeout)
        with self._lock:
            self._expected_command = None
            return self._result if got else None
 
 
# Globale postbus-instantie (naast mav_connection / mav_lock)
ack_mailbox = AckMailbox()

# ============================================
# COORDINATE LOG STORAGE (JSON op disk)
# ============================================
#
# Het gelogde-coordinaten-bestand is de "source of truth" voor alle
# entries. Frontend leest het via GET /api/log bij page-load en
# wijzigt het via POST/PUT/DELETE endpoints.
#
# Bestandsformaat: lijst van entry-objects, top-level array.
#   [
#     {"id": "abc123", "lat": 50.7, "lon": 4.3, "alt": 0,
#      "time": "15:44:22", "date": "2026-05-20T13:44:22.829Z",
#      "source": "manueel"|"drone",
#      "status": "gemeld"|"wordt_onderzocht"|"waargenomen"|"bestreden"|"vals_alarm"|"",
#      "notes": ""},
#     ...
#   ]
#
# ID is een server-side gegenereerde unieke string (timestamp+random).
# Frontend gebruikt deze voor PUT/DELETE in plaats van array-index zodat
# delete + concurrent edit niet de verkeerde entry raakt.

def ensure_data_dir():
    """Maak data/ aan als hij niet bestaat (bv. eerste run)."""
    os.makedirs(DATA_DIR, exist_ok=True)


def load_log():
    """
    Lees coord-log uit JSON-bestand. Returns lege lijst als bestand niet
    bestaat of corrupt is. Thread-safe via log_lock.
    """
    with log_lock:
        if not os.path.exists(LOG_FILE):
            return []
        try:
            with open(LOG_FILE, 'r', encoding='utf-8') as f:
                data = json.load(f)
            if isinstance(data, list):
                return data
            print(f"!! coord-log.json bevat geen lijst, return leeg")
            return []
        except (json.JSONDecodeError, IOError) as e:
            print(f"!! coord-log.json lezen mislukt: {e}, return leeg")
            return []


def save_log(entries):
    """
    Schrijf coord-log atomisch naar disk: write naar .tmp file, dan
    os.replace() naar definitieve naam. os.replace is atomair op POSIX
    zodat het bestand nooit half-geschreven op disk staat bij een crash.
    Thread-safe via log_lock.
    """
    with log_lock:
        ensure_data_dir()
        tmp_path = LOG_FILE + '.tmp'
        try:
            with open(tmp_path, 'w', encoding='utf-8') as f:
                json.dump(entries, f, ensure_ascii=False, indent=2)
            os.replace(tmp_path, LOG_FILE)
            return True
        except IOError as e:
            print(f"!! coord-log.json schrijven mislukt: {e}")
            return False


def generate_entry_id():
    """
    Genereer een unieke string-ID voor een nieuwe entry.
    Gebruikt timestamp (ms) + 4 random hex chars zodat collisions
    onmogelijk zijn bij realistische gebruiks-rate.
    """
    import secrets
    timestamp_ms = int(time.time() * 1000)
    random_hex = secrets.token_hex(2)
    return f"{timestamp_ms}-{random_hex}"


def reverse_geocode(lat, lon, timeout=2.0):
    """
    Roep Nominatim reverse-geocode API aan om lat/lon naar leesbaar adres
    te converteren. Returns een string in formaat "Gemeente, Straat Nummer"
    of None bij geen internet / timeout / geen resultaat.

    Voorbeelden:
        50.7686, 4.2700  ->  "Vlezenbeek, Kerkstraat 5"
        50.7100, 4.3500  ->  "Anderlecht, Nijverheidskaai 12"
        50.7500, 4.2000  ->  "Sint-Pieters-Leeuw" (alleen gemeente bij weide)
        midden in zee    ->  None

    Nominatim usage policy:
        - User-Agent vereist met identificatie
        - Max 1 request per seconde (rate-limit), wij hebben hier geen
          concurrent requests dus geen extra throttling nodig
        - Bij commercieel gebruik: eigen Nominatim instance overwegen

    Gebruikt twee seconden timeout zodat een trage Nominatim of geen
    internet de log-opslag niet onnodig lang laat hangen.
    """
    import urllib.request
    import urllib.error

    url = (
        f"https://nominatim.openstreetmap.org/reverse"
        f"?format=jsonv2"
        f"&lat={lat}"
        f"&lon={lon}"
        f"&zoom=18"            # straat-niveau detail
        f"&addressdetails=1"   # opgesplitste address-velden in response
        f"&accept-language=nl"
    )
    req = urllib.request.Request(url, headers={'User-Agent': TILE_USER_AGENT})

    try:
        with urllib.request.urlopen(req, timeout=timeout) as response:
            data = json.loads(response.read())
    except (urllib.error.URLError, TimeoutError, json.JSONDecodeError, OSError) as e:
        print(f"[reverse_geocode] failed for {lat},{lon}: {e}")
        return None

    addr = data.get('address', {}) or {}

    # Gemeente-naam: probeer in volgorde van specifiek naar algemeen.
    # OSM gebruikt verschillende velden afhankelijk van land/regio:
    #   - village    voor dorpen
    #   - town       voor kleinere steden
    #   - city       voor steden
    #   - municipality is de officiele gemeente-grens (overkoepelend)
    # We pakken de eerste die niet leeg is.
    gemeente = (
        addr.get('village')
        or addr.get('town')
        or addr.get('city')
        or addr.get('municipality')
        or addr.get('suburb')
        or addr.get('county')
    )

    # Straat + huisnummer combineren
    straat = addr.get('road') or addr.get('pedestrian') or addr.get('footway')
    nummer = addr.get('house_number')

    # Bouw output-string in formaat "Gemeente, Straat Nummer"
    if gemeente and straat and nummer:
        return f"{gemeente}, {straat} {nummer}"
    if gemeente and straat:
        return f"{gemeente}, {straat}"
    if gemeente:
        return gemeente
    if straat and nummer:
        return f"{straat} {nummer}"
    if straat:
        return straat

    return None


def ensure_addresses_filled(entries, save_after=True):
    """
    Loop over entries en haal voor elke entry zonder adres alsnog
    via Nominatim het adres op. Slaat de aangevulde lijst terug naar
    disk zodat we deze ophaal-poging niet bij elke export opnieuw doen.

    Nominatim rate-limit: 1 request/sec. Bij 30 entries zonder adres
    duurt dit ~30 sec. Voor offline veld-batch acceptabel.

    Bij geen internet voor de eerste entry: stoppen we vroeg
    (geen zin om 30x dezelfde error te krijgen) en geven terug wat
    we hebben.

    Returns (entries, n_filled): bijgewerkte lijst + hoeveel adressen
    we deze keer extra hebben opgehaald.
    """
    n_filled = 0
    n_failed_in_a_row = 0

    for entry in entries:
        if entry.get('address'):
            continue  # al ingevuld, skip

        # Stop vroeg als we drie keer op rij faalden — waarschijnlijk
        # geen internet, geen zin om door te gaan
        if n_failed_in_a_row >= 3:
            print(f"[export] 3x op rij geen Nominatim-response, "
                  f"stop met aanvullen")
            break

        lat = entry.get('lat')
        lon = entry.get('lon')
        if lat is None or lon is None:
            continue

        address = reverse_geocode(lat, lon)
        if address:
            entry['address'] = address
            n_filled += 1
            n_failed_in_a_row = 0
            # Nominatim rate-limit: minstens 1 sec tussen requests
            time.sleep(1.1)
        else:
            n_failed_in_a_row += 1

    if n_filled > 0 and save_after:
        save_log(entries)
        print(f"[export] {n_filled} adressen alsnog opgehaald + gecached")

    return entries, n_filled
    # ============================================
# TILE CACHE (offline maps)
# ============================================
#
# Tiles worden lokaal opgeslagen onder data/tiles/<source>/<z>/<x>/<y>.png.
# De serve_tile route hierboven probeert eerst lokaal, valt terug op
# internet wanneer een tile niet gecached is.
#
# Atomic writes: download naar .tmp, dan os.replace zodat een crashed
# download geen corrupte PNG achterlaat.

def tile_cache_path(source, z, x, y):
    """Bouw het filesystem-pad voor een tile in onze cache."""
    return os.path.join(TILE_CACHE_DIR, source, str(z), str(x), f"{y}.png")


def fetch_tile_from_internet(source, z, x, y):
    """
    Haal een tile op bij de externe tile-server. Returns (bytes, content_type)
    of (None, None) bij fout (netwerk, 404, blokkering).

    Slaat het resultaat NIET zelf op — dat doet de aanroeper, zodat
    fetch en cache-write apart te debuggen zijn.
    """
    import urllib.request
    import urllib.error

    if source not in TILE_SOURCES:
        return None, None

    url = TILE_SOURCES[source].format(z=z, x=x, y=y)
    req = urllib.request.Request(url, headers={'User-Agent': TILE_USER_AGENT})

    try:
        with urllib.request.urlopen(req, timeout=5) as response:
            data = response.read()
            content_type = response.headers.get('Content-Type', 'image/png')
            return data, content_type
    except urllib.error.URLError as e:
        # Netwerk down, geen internet, of tile server unreachable
        print(f"[tiles] fetch failed {source}/{z}/{x}/{y}: {e}")
        return None, None
    except Exception as e:
        print(f"[tiles] unexpected error {source}/{z}/{x}/{y}: {e}")
        return None, None


def save_tile_to_cache(source, z, x, y, data):
    """
    Schrijf een tile atomisch naar de cache. Maakt parent-directories aan
    indien nodig. Returns True bij succes.
    """
    path = tile_cache_path(source, z, x, y)
    try:
        os.makedirs(os.path.dirname(path), exist_ok=True)
        tmp_path = path + '.tmp'
        with open(tmp_path, 'wb') as f:
            f.write(data)
        os.replace(tmp_path, path)
        return True
    except IOError as e:
        print(f"[tiles] save failed {source}/{z}/{x}/{y}: {e}")
        return False


        # --- Tile berekeningen (slippy-map / mercator) ---
# Deze functies zijn een copy van prefetch_tiles.py zodat de Flask-route
# zelfstandig kan rekenen zonder dat script te moeten importeren. Bij een
# latere refactor zou tile_utils.py een gedeelde module kunnen worden.

def deg2num(lat_deg, lon_deg, zoom):
    """Lat/lon -> tile (x, y) op een gegeven zoom (Slippy-map convention)."""
    lat_rad = math.radians(lat_deg)
    n = 2.0 ** zoom
    xtile = int((lon_deg + 180.0) / 360.0 * n)
    ytile = int((1.0 - math.asinh(math.tan(lat_rad)) / math.pi) / 2.0 * n)
    return xtile, ytile


def tiles_for_area(lat, lon, radius_km, zoom):
    """Bereken alle (x, y) tile-coords binnen een radius op een zoom-level."""
    lat_offset = radius_km / 111.0
    lon_offset = radius_km / (111.0 * math.cos(math.radians(lat)))
    lat_min, lat_max = lat - lat_offset, lat + lat_offset
    lon_min, lon_max = lon - lon_offset, lon + lon_offset

    # y is omgekeerd in slippy-map (noord = ymin)
    x_min, y_max = deg2num(lat_min, lon_min, zoom)
    x_max, y_min = deg2num(lat_max, lon_max, zoom)

    tiles = []
    for x in range(x_min, x_max + 1):
        for y in range(y_min, y_max + 1):
            tiles.append((x, y))
    return tiles


# --- Prefetch background state ---
# Eén lopende prefetch tegelijk. Frontend pollt /api/tiles/prefetch/status
# voor voortgang. State leeft in geheugen, gaat verloren bij service-restart
# (acceptabel: bij restart zou prefetch sowieso interrupt zijn).

prefetch_state = {
    'running': False,
    'started_at': None,
    'total': 0,
    'done': 0,
    'success': 0,
    'fail': 0,
    'message': '',
}
prefetch_lock = threading.Lock()


def run_prefetch(lat, lon, radius_km, zoom_min, zoom_max, sources):
    """
    Background-thread functie die tiles ophaalt via fetch_tile_from_internet
    en in de cache opslaat. Updatet prefetch_state zodat de frontend kan pollen.
    """
    # Bouw eerst de volledige werklijst zodat 'total' correct is
    work = []  # (source, z, x, y)
    for source in sources:
        if source not in TILE_SOURCES:
            continue
        for z in range(zoom_min, zoom_max + 1):
            for (x, y) in tiles_for_area(lat, lon, radius_km, z):
                work.append((source, z, x, y))

    with prefetch_lock:
        prefetch_state['total'] = len(work)
        prefetch_state['done'] = 0
        prefetch_state['success'] = 0
        prefetch_state['fail'] = 0
        prefetch_state['message'] = f'Bezig: {len(work)} tiles'

    for (source, z, x, y) in work:
        # Skip als al gecached — bespaart bandwidth bij herhaalde prefetch
        path = tile_cache_path(source, z, x, y)
        if os.path.exists(path):
            with prefetch_lock:
                prefetch_state['done'] += 1
                prefetch_state['success'] += 1
            continue

        data, _ = fetch_tile_from_internet(source, z, x, y)
        if data is not None:
            save_tile_to_cache(source, z, x, y, data)
            with prefetch_lock:
                prefetch_state['success'] += 1
        else:
            with prefetch_lock:
                prefetch_state['fail'] += 1

        with prefetch_lock:
            prefetch_state['done'] += 1

    with prefetch_lock:
        prefetch_state['running'] = False
        prefetch_state['message'] = (
            f'Klaar: {prefetch_state["success"]} OK, '
            f'{prefetch_state["fail"]} fail'
        )


# ============================================
# SIGNAL SOURCE INTERFACE (abstract)
# ============================================
class SignalSource:
    """
    Abstracte klasse. Subclasses implementeren:
      - open()                : init hardware
      - close()               : cleanup
      - measure_once() -> dB  : 1 single narrow-band meting in dB
      - describe() -> str     : voor logging

    Voor LoRa ontvanger kan measure_once() simpelweg de laatste
    gedecodeerde packet RSSI returnen, met detected flag in status.
    """
    def open(self): raise NotImplementedError
    def close(self): pass
    def measure_once(self): raise NotImplementedError
    def describe(self): return "unknown"


# ============================================
# RTL-SDR SIGNAL SOURCE
# ============================================
class RtlSdrSource(SignalSource):
    def __init__(self):
        self.sdr = None

    def open(self):
        from rtlsdr import RtlSdr
        while self.sdr is None:
            try:
                print("RTL-SDR openen...")
                sdr = RtlSdr()
                sdr.sample_rate = SAMPLE_RATE
                sdr.center_freq = CENTER_FREQ
                sdr.gain = GAIN
                sdr.read_samples(NUM_SAMPLES)  # warmup
                self.sdr = sdr
                print(f"RTL-SDR OK: {SAMPLE_RATE/1e6:.2f} MSPS @ {CENTER_FREQ/1e6:.3f} MHz, gain {GAIN} dB")
            except Exception as e:
                print(f"RTL-SDR fout: {e}. Retry over 3s...")
                time.sleep(3)

    def close(self):
        if self.sdr:
            try: self.sdr.close()
            except Exception: pass
            self.sdr = None

    def measure_once(self):
        try:
            samples = self.sdr.read_samples(NUM_SAMPLES)
        except Exception as e:
            print(f"SDR read error: {e}, reopen...")
            self.close()
            self.open()
            return None

        # FFT, shift zodat DC in midden zit
        fft = np.fft.fftshift(np.fft.fft(samples))
        power = np.abs(fft) ** 2

        bin_hz = SAMPLE_RATE / len(samples)
        center_bin = len(samples) // 2
        half_bins = int((DETECT_BANDWIDTH / 2) / bin_hz)

        # DC-spike vermijden (±2 bins rond center)
        dc_gap = 2
        left = power[center_bin - half_bins : center_bin - dc_gap]
        right = power[center_bin + dc_gap : center_bin + half_bins]
        narrow = np.concatenate([left, right])
        if len(narrow) == 0:
            return -100.0

        peak = float(np.max(narrow))
        return 10 * np.log10(peak + 1e-12)

    def describe(self):
        return f"RTL-SDR @ {CENTER_FREQ/1e6:.3f} MHz, BW {DETECT_BANDWIDTH/1e3:.0f} kHz"


# ============================================
# LORA SIGNAL SOURCE (Ra-01 / SX1278 via SPI)
# ============================================
class LoRaSource(SignalSource):
    """
    Echte Ra-01 / SX1278 LoRa ontvanger via SPI.

    Achtergrond-thread leest packets via polling van het IRQ-register
    (interrupt-callback werkt niet betrouwbaar met rpi-lgpio op Pi OS
    Bookworm). Bij elk ontvangen 'HT,<id>,<count>' packet bewaren we
    RSSI/SNR en metadata in instance-variabelen.

    measure_once() returns:
      - laatst-gehoorde RSSI als binnen 3s een packet ontvangen werd
      - -120.0 (floor) anders, zodat 'geen signaal' duidelijk zichtbaar is
        in de signal-bar (ver onder baseline) zonder dat we de bestaande
        signal-display logica hoeven aanpassen.

    Update status dict velden:
      - lora_packet_count       totaal aantal succesvol ontvangen packets
      - lora_last_tracker_id    ID uit laatste packet
      - lora_last_seen_sec      seconden sinds laatste packet (-1 = nooit)
      - lora_snr                SNR van laatste packet in dB (echte waarde)
    """

    POLL_INTERVAL = 0.05    # 50 ms — fijn genoeg voor 1 Hz beacon
    SILENCE_FLOOR = -120.0  # dBm — value bij stilte >3s
    SILENCE_TIMEOUT = 3.0   # sec zonder packet -> floor

    def __init__(self):
        self._lora = None
        self._board = None
        self._rx_thread = None
        self._stop = threading.Event()

        # Latest packet state — alleen update bij ontvangst, dus thread-safe
        # om uit te lezen (atomic float/int writes in Python).
        self._last_rssi = self.SILENCE_FLOOR
        self._last_snr = 0.0
        self._last_seen = 0.0  # time.time() bij laatste packet, 0 = nooit
        self._tracker_id = 0
        self._packet_count = 0

    def open(self):
        """Init Ra-01, start RX-thread."""
        try:
            from SX127x.LoRa import LoRa, MODE
            from SX127x.board_config import BOARD
        except ImportError as e:
            raise RuntimeError(f"pyLoRa niet beschikbaar: {e}")

        print("LoRa Ra-01 openen...")
        BOARD.setup()
        self._board = BOARD

        class _Lora(LoRa):
            def __init__(self):
                super().__init__(verbose=False)

        lora = _Lora()
        lora.set_mode(MODE.SLEEP)
        lora.set_dio_mapping([0] * 6)

        # Match beacon config exact (zie beacon.ino):
        # 433 MHz, BW 125 kHz, SF 7, CR 4/5, CRC on, sync word default
        lora.set_freq(433.0)
        lora.set_bw(7)            # BW index 7 = 125 kHz
        lora.set_spreading_factor(7)
        lora.set_coding_rate(1)   # 1 = 4/5
        lora.set_rx_crc(True)
        lora.set_pa_config(pa_select=1)
        lora.set_lna_gain(1)      # G1 = max LNA gain (1=max, 6=min)
        lora.set_implicit_header_mode(False)

        # Start continuous RX
        lora.reset_ptr_rx()
        lora.set_mode(MODE.RXCONT)
        self._lora = lora

        # Verify chip — RegVersion moet 0x12 zijn
        version = lora.get_version()
        if version != 0x12:
            raise RuntimeError(
                f"SX1278 RegVersion = 0x{version:02x}, verwacht 0x12. "
                f"Check bedrading (NSS/MISO/MOSI/SCK/RST/DIO0)."
            )
        print(f"LoRa Ra-01 OK: RegVersion=0x{version:02x}, "
              f"433.0 MHz, BW 125 kHz, SF 7")

        # Start RX-thread
        self._stop.clear()
        self._rx_thread = threading.Thread(
            target=self._rx_loop, daemon=True, name='lora-rx'
        )
        self._rx_thread.start()

    def _rx_loop(self):
        """
        Polling-loop op IRQ-register. Voor elke RxDone: decode payload,
        update _last_rssi / _last_snr / _tracker_id / _packet_count.
        """
        global status
        from SX127x.LoRa import MODE

        while not self._stop.is_set():
            time.sleep(self.POLL_INTERVAL)
            try:
                flags = self._lora.get_irq_flags()
            except Exception as e:
                print(f"!! LoRa IRQ read fout: {e}")
                continue

            if not flags.get('rx_done', 0):
                continue

            # Packet ontvangen
            try:
                crc_error = flags.get('crc_error', 0)
                payload_raw = self._lora.read_payload(nocheck=True)
                rssi = self._lora.get_pkt_rssi_value()
                snr_value = self._lora.get_pkt_snr_value()

                # Clear flags + restart RX voor volgende packet
                # Wis alle relevante IRQ flags na packet-handling
                self._lora.clear_irq_flags(
                    RxDone=1,
                    PayloadCrcError=1,
                    ValidHeader=1,
                )
                self._lora.reset_ptr_rx()
                self._lora.set_mode(MODE.RXCONT)
            except Exception as e:
                print(f"!! LoRa packet read fout: {e}")
                continue

            if crc_error:
                # CRC fout — packet onbruikbaar, skip
                continue

            # Decode ASCII payload "HT,<id>,<count>"
            try:
                text = bytes(payload_raw).decode('ascii', errors='strict')
            except UnicodeDecodeError:
                continue   # niet-ASCII, niet onze beacon

            parts = text.split(',')
            if len(parts) != 3 or parts[0] != 'HT':
                continue   # niet ons protocol

            try:
                tracker_id = int(parts[1])
                # tx_count interesseert ons hier niet, maar valideert format
                int(parts[2])
            except ValueError:
                continue

            # pyLoRa's get_pkt_snr_value() past de kwart-dB-schaling al toe; de
            # waarde is dus geen ruwe registerbyte. De oude heuristiek
            # (`snr_raw if abs(...) <= 20 else snr_raw / 4.0`) deelde in het
            # grote bereik een tweede keer en was daarmee fout.
            # Zie SX127x/LoRa.py -> get_pkt_snr_value().
            snr_db = snr_value

            # Update state — alleen primitieve types, atomic in CPython
            self._last_rssi = float(rssi)
            self._last_snr = float(snr_db)
            self._last_seen = time.time()
            self._tracker_id = tracker_id
            self._packet_count += 1

            # Update status dict zodat dashboard de extra info ziet
            status['lora_packet_count'] = self._packet_count
            status['lora_last_tracker_id'] = tracker_id
            status['lora_snr'] = round(snr_db, 1)
            status['lora_last_seen_sec'] = 0

    def measure_once(self):
        """
        Returns laatste RSSI in dBm, of -120 dBm bij stilte >3s.

        NB: bestaande signal_loop roept measure_once() 50x per cycle aan
        (peak-hold). Voor LoRa heeft dat geen nut omdat we al de echte
        packet-RSSI hebben — maar het is harmless want we returnen gewoon
        dezelfde laatste waarde. Signal_loop heeft een speciaal pad voor
        LoRa zodat we maar 1x per cycle de waarde lezen.
        """
        global status
        now = time.time()
        if self._last_seen == 0:
            # Nog nooit een packet ontvangen
            status['lora_last_seen_sec'] = -1
            status['lora_snr'] = 0.0      # geen signaal -> geen getrouwheid
            return self.SILENCE_FLOOR

        age = now - self._last_seen
        status['lora_last_seen_sec'] = round(age, 1)

        if age > self.SILENCE_TIMEOUT:
            status['lora_snr'] = 0.0      # geen signaal -> geen getrouwheid
            return self.SILENCE_FLOOR

        return self._last_rssi

    def close(self):
        """Stop RX-thread en cleanup."""
        self._stop.set()
        if self._rx_thread and self._rx_thread.is_alive():
            self._rx_thread.join(timeout=2)
        if self._lora is not None:
            try:
                from SX127x.LoRa import MODE
                self._lora.set_mode(MODE.SLEEP)
            except Exception:
                pass
            self._lora = None
        if self._board is not None:
            try:
                self._board.teardown()
            except Exception:
                pass
            self._board = None

    def describe(self):
        return "LoRa Ra-01 SX1278 @ 433.0 MHz, BW 125 kHz, SF 7"


def make_signal_source():
    """
    Bouw de SignalSource volgens SIGNAL_SOURCE config.

    Geen fallback meer: als 'lora' faalt willen we de error duidelijk
    zien (Pi-side log + dashboard 'signaal -120 dBm'), niet stilletjes
    terugvallen op RTL-SDR. Dat verhult hardware-problemen.
    """
    if SIGNAL_SOURCE == 'lora':
        src = LoRaSource()
        src.open()
        return src
    if SIGNAL_SOURCE == 'rtlsdr':
        src = RtlSdrSource()
        src.open()
        return src
    raise ValueError(f"Onbekende SIGNAL_SOURCE: {SIGNAL_SOURCE!r}")


# ============================================
# SIGNAL LOOP (peak-hold + vaste baseline)
# ============================================
def signal_loop():
    """
    Centrale signal-loop met twee paden:
      - RTL-SDR: peak-hold + dynamische baseline (oude flow, ongewijzigd)
      - LoRa:    direct laatste packet-RSSI + packet-age detectie

    Voor LoRa zijn signal_delta/baseline geen betekenisvolle concepten —
    we hebben echte packet-RSSI in dBm. signal_active is true zolang er
    binnen 3s een packet binnenkwam.
    """
    global status, baseline_reset_requested

    source = make_signal_source()
    status['signal_source'] = source.describe()
    print(f"Signal source: {source.describe()}")

    # --- LoRa pad: geen baseline meting, direct in loop ---
    if SIGNAL_SOURCE == 'lora':
        status['baseline'] = -120.0   # placeholder, niet gebruikt
        socketio.emit('baseline_status', {
            'measuring': False, 'baseline': -120.0
        })
        print("LoRa modus: baseline-meting overgeslagen "
              "(LoRa gebruikt packet-based detectie)")

        while True:
            try:
                rssi = source.measure_once()
                last_seen = status.get('lora_last_seen_sec', -1)

                # Detectie: packet binnen 3s = signaal actief
                active = (last_seen >= 0) and (last_seen < 3.0)

                # Hou signal_power op laatste echte RSSI, ook tijdens
                # 'stilte' tussen packets. Frontend toont dan een stabiele
                # waarde in plaats van te flikkeren tussen -94 en -120.
                if active:
                    status['signal_power'] = round(rssi, 1)
                else:
                    status['signal_power'] = -120.0

                status['signal_delta'] = 0.0      # niet betekenisvol voor LoRa
                status['signal_detected'] = active

                socketio.emit('status_update', status)
                time.sleep(MEASURE_INTERVAL)

            except Exception as e:
                print(f"Signal loop error (LoRa): {e}")
                time.sleep(1)

    # --- RTL-SDR pad: bestaande peak-hold + baseline flow ---
    def do_baseline():
        print("Baseline meting...")
        socketio.emit('baseline_status', {'measuring': True})
        readings = []
        for i in range(BASELINE_NUM_READINGS):
            peaks = []
            for _ in range(PEAK_HOLD_SAMPLES):
                v = source.measure_once()
                if v is not None:
                    peaks.append(v)
            if peaks:
                readings.append(max(peaks))
                print(f"  Baseline {i+1}/{BASELINE_NUM_READINGS}: {readings[-1]:.1f} dB")
        baseline = sum(readings) / len(readings) if readings else -100
        print(f"Baseline: {baseline:.1f} dB (vast)")
        socketio.emit('baseline_status', {
            'measuring': False, 'baseline': round(baseline, 1)
        })
        return baseline

    baseline = do_baseline()
    status['baseline'] = round(baseline, 1)

    while True:
        try:
            if baseline_reset_requested:
                baseline_reset_requested = False
                baseline = do_baseline()
                status['baseline'] = round(baseline, 1)

            peaks = []
            for _ in range(PEAK_HOLD_SAMPLES):
                v = source.measure_once()
                if v is not None:
                    peaks.append(v)

            if not peaks:
                time.sleep(MEASURE_INTERVAL)
                continue

            peak_power = max(peaks)
            delta = peak_power - baseline
            detected = delta > THRESHOLD

            status['signal_power'] = round(peak_power, 1)
            status['signal_delta'] = round(delta, 1)
            status['signal_detected'] = detected

            socketio.emit('status_update', status)
            time.sleep(MEASURE_INTERVAL)

        except Exception as e:
            print(f"Signal loop error: {e}")
            time.sleep(1)


# ============================================
# WIFI RSSI (ongewijzigd)
# ============================================
def read_wifi_status():
    try:
        result = subprocess.run(
            ['iw', 'dev', WIFI_HOTSPOT_IFACE, 'station', 'dump'],
            capture_output=True, text=True, timeout=3
        )
        output = result.stdout
        if not output.strip():
            return {'connected': False, 'rssi': -100, 'quality': 0, 'clients': 0}
        clients = len(re.findall(r'^Station ', output, re.MULTILINE))
        signals = [int(m) for m in re.findall(r'signal:\s*(-?\d+)', output)]
        if not signals:
            return {'connected': clients > 0, 'rssi': -100, 'quality': 0, 'clients': clients}
        best_rssi = max(signals)
        quality = max(0, min(100, int((best_rssi + 90) * 100 / 60)))
        return {'connected': True, 'rssi': best_rssi, 'quality': quality, 'clients': clients}
    except FileNotFoundError:
        return {'connected': False, 'rssi': -100, 'quality': 0, 'clients': 0}
    except Exception as e:
        print(f"WiFi status error: {e}")
        return {'connected': False, 'rssi': -100, 'quality': 0, 'clients': 0}


def wifi_loop():
    global status
    while True:
        try:
            wifi = read_wifi_status()
            status['wifi_connected'] = wifi['connected']
            status['wifi_rssi'] = wifi['rssi']
            status['wifi_quality'] = wifi['quality']
            status['wifi_clients'] = wifi['clients']
            time.sleep(2)
        except Exception as e:
            print(f"WiFi loop error: {e}")
            time.sleep(2)


# ============================================
# MAVLINK (ongewijzigd)
# ============================================
def mavlink_loop():
    global status, mav_connection
    try:
        from pymavlink import mavutil
    except ImportError:
        print("!! pymavlink niet geinstalleerd")
        return
 
    COPTER_MODES = {
        0: 'STABILIZE', 1: 'ACRO', 2: 'ALT_HOLD', 3: 'AUTO',
        4: 'GUIDED', 5: 'LOITER', 6: 'RTL', 7: 'CIRCLE',
        9: 'LAND', 11: 'DRIFT', 13: 'SPORT', 14: 'FLIP',
        15: 'AUTOTUNE', 16: 'POSHOLD', 17: 'BRAKE', 18: 'THROW',
        19: 'AVOID_ADSB', 20: 'GUIDED_NOGPS', 21: 'SMART_RTL',
        22: 'FLOWHOLD', 23: 'FOLLOW', 24: 'ZIGZAG', 25: 'SYSTEMID',
        26: 'AUTOROTATE', 27: 'AUTO_RTL'
    }
    last_radio_status = 0
 
    while True:
        try:
            print(f"MAVLink verbinden met {MAVLINK_DEVICE} @ {MAVLINK_BAUD}...")
            mav = mavutil.mavlink_connection(MAVLINK_DEVICE, baud=MAVLINK_BAUD)
            mav.wait_heartbeat(timeout=10)
            print(f"MAVLink OK! System {mav.target_system}, Component {mav.target_component}")
            status['pixhawk_connected'] = True
            with mav_lock:
                mav_connection = mav
 
            mav.mav.request_data_stream_send(
                mav.target_system, mav.target_component,
                mavutil.mavlink.MAV_DATA_STREAM_ALL, 4, 1
            )
 
            while True:
                msg = mav.recv_match(blocking=True, timeout=5)
                if msg is None:
                    print("MAVLink timeout, herverbinden...")
                    status['pixhawk_connected'] = False
                    with mav_lock:
                        mav_connection = None
                    break
                msg_type = msg.get_type()
                now = time.time()
 
                # ---- NIEUW: COMMAND_ACK opvangen en in de postbus leggen ----
                # Dit is de kern van de fix: de centrale lus is de ENIGE die
                # recv_match doet, en distribueert acks naar wachtende handlers.
                if msg_type == 'COMMAND_ACK':
                    ack_mailbox.deliver(msg.command, msg.result)
                    # We gaan door — een COMMAND_ACK bevat geen telemetrie.
                    continue
 
                if msg_type == 'HEARTBEAT':
                    status['pixhawk_connected'] = True
                    status['armed'] = bool(msg.base_mode & mavutil.mavlink.MAV_MODE_FLAG_SAFETY_ARMED)
                    status['flight_mode'] = COPTER_MODES.get(msg.custom_mode, f'MODE_{msg.custom_mode}')
                elif msg_type == 'SYS_STATUS':
                    status['battery_voltage'] = round(msg.voltage_battery / 1000.0, 2)
                    if msg.battery_remaining >= 0:
                        status['battery_percent'] = msg.battery_remaining
                elif msg_type == 'BATTERY_STATUS':
                    if msg.voltages[0] != 65535:
                        status['battery_voltage'] = round(msg.voltages[0] / 1000.0, 2)
                    if msg.battery_remaining >= 0:
                        status['battery_percent'] = msg.battery_remaining
                elif msg_type == 'GPS_RAW_INT':
                    status['gps_fix'] = msg.fix_type >= 3
                    status['gps_satellites'] = msg.satellites_visible
                elif msg_type == 'GLOBAL_POSITION_INT':
                    status['gps_lat'] = msg.lat / 1e7
                    status['gps_lon'] = msg.lon / 1e7
                    status['altitude'] = round(msg.relative_alt / 1000.0, 2)
                    if msg.hdg != 65535:
                        status['heading'] = round(msg.hdg / 100.0, 1)
                elif msg_type in ('RADIO_STATUS', 'RADIO'):
                    last_radio_status = now
                    status['telem_connected'] = True
                    status['telem_rssi_local'] = round(msg.rssi / 1.9 - 127, 1)
                    status['telem_rssi_remote'] = round(msg.remrssi / 1.9 - 127, 1)
                    status['telem_noise_local'] = round(msg.noise / 1.9 - 127, 1)
                    status['telem_noise_remote'] = round(msg.remnoise / 1.9 - 127, 1)
                    status['telem_txbuf'] = msg.txbuf
                    worst_snr = min(msg.rssi - msg.noise, msg.remrssi - msg.remnoise)
                    status['telem_quality'] = max(0, min(100, int(worst_snr * 100 / 50)))
 
                if now - last_radio_status > 5:
                    status['telem_connected'] = False
        except Exception as e:
            print(f"MAVLink error: {e}")
            status['pixhawk_connected'] = False
            status['telem_connected'] = False
            time.sleep(5)

# ============================================
# THERMAL CAMERA LOOP (Pimoroni MLX90640)
# ============================================
#
# Achtergrondthread die continu thermische frames leest via I2C en
# ze als Socket.io 'thermal_frame' event naar het dashboard pusht.
#
# Frame-formaat: 32x24 = 768 floats in °C. Wordt verstuurd als platte
# lijst; frontend reconstrueert 32-kolom layout met index = h*32 + w.
#
# Skip-frames zijn normaal bij hogere refresh rates — de MLX90640
# stuurt elk frame in twee subframes en bij timing-issues moet de
# library hertest. We loggen ze maar laten de loop doorgaan.

def thermal_loop():
    global status

    try:
        import board
        import busio
        import adafruit_mlx90640
    except ImportError as e:
        print(f"!! MLX90640 libraries niet beschikbaar: {e}")
        print("!! Installeer: pip3 install adafruit-circuitpython-mlx90640 --break-system-packages")
        return

    # Refresh-rate enum mapping
    refresh_rates = {
        2: adafruit_mlx90640.RefreshRate.REFRESH_2_HZ,
        4: adafruit_mlx90640.RefreshRate.REFRESH_4_HZ,
        8: adafruit_mlx90640.RefreshRate.REFRESH_8_HZ,
        16: adafruit_mlx90640.RefreshRate.REFRESH_16_HZ,
        32: adafruit_mlx90640.RefreshRate.REFRESH_32_HZ,
    }
    refresh_enum = refresh_rates.get(THERMAL_REFRESH_HZ,
                                      adafruit_mlx90640.RefreshRate.REFRESH_8_HZ)

    # Verbinding opzetten — retry-loop zodat een tijdelijke I2C-glitch
    # niet de hele thread doodt
    mlx = None
    while mlx is None:
        try:
            print("Thermal camera openen...")
            i2c = busio.I2C(board.SCL, board.SDA, frequency=800000)
            mlx = adafruit_mlx90640.MLX90640(i2c)
            mlx.refresh_rate = refresh_enum
            print(f"Thermal camera OK: serial {[hex(i) for i in mlx.serial_number]}, "
                  f"refresh {THERMAL_REFRESH_HZ} Hz")
            status['thermal_connected'] = True
        except (ValueError, OSError) as e:
            print(f"Thermal camera fout: {e}. Retry over 5s...")
            time.sleep(5)

    # Frame buffer wordt door getFrame in-place gevuld
    frame = [0.0] * 768
    last_emit = 0.0
    fail_count = 0

    # Voor FPS-meting: rolling window van frame timestamps
    frame_times = []

    while True:
        try:
            mlx.getFrame(frame)
            fail_count = 0  # reset bij succes
            
            # Bewaar laatste frame voor baseline-capture handler.
            # Module-scope zodat de Socket.io handlers erbij kunnen.
            global _last_thermal_frame
            _last_thermal_frame = list(frame)

            now = time.time()
            frame_times.append(now)
            # Houd alleen laatste 2 seconden voor FPS-berekening
            frame_times = [t for t in frame_times if now - t < 2.0]
            fps = len(frame_times) / 2.0 if len(frame_times) > 1 else 0.0

            # Rate-limit emits: ook al kan camera 8 FPS, browser hoeft
            # niet zo vaak te updaten. THERMAL_EMIT_INTERVAL bepaalt cadence.
            if now - last_emit >= THERMAL_EMIT_INTERVAL:
                last_emit = now

                # Statistieken berekenen (single-pass voor performance)
                mn = mx = frame[0]
                total = 0.0
                for v in frame:
                    if v < mn: mn = v
                    if v > mx: mx = v
                    total += v
                avg = total / 768

                status['thermal_min'] = round(mn, 1)
                status['thermal_max'] = round(mx, 1)
                status['thermal_avg'] = round(avg, 1)
                status['thermal_fps'] = round(fps, 1)

                # Emit frame als aparte event om status_update klein te houden.
                # Frame is een lijst van 768 floats met 1 decimaal afgerond
                # zodat JSON-grootte ~5 KB blijft i.p.v. ~15 KB.
                #
                # Baseline wordt meegestuurd in elke emit zodat de frontend altijd
                # weet of detectie-modus mogelijk is. None als nog niet ingesteld.
                socketio.emit('thermal_frame', {
                    'data': [round(v, 1) for v in frame],
                    'min': status['thermal_min'],
                    'max': status['thermal_max'],
                    'avg': status['thermal_avg'],
                    'fps': status['thermal_fps'],
                    'baseline': thermal_baseline,
                })

        except (ValueError, RuntimeError) as e:
            # "Too many retries" of "Frame data error" — normaal bij snelle
            # refresh-rates. Tel ze maar paniekeer niet.
            fail_count += 1
            if fail_count > 50:
                print(f"!! Veel skipped frames ({fail_count}), camera mogelijk losgeraakt")
                status['thermal_connected'] = False
                # Probeer opnieuw te verbinden
                fail_count = 0
                mlx = None
                while mlx is None:
                    try:
                        time.sleep(2)
                        i2c = busio.I2C(board.SCL, board.SDA, frequency=800000)
                        mlx = adafruit_mlx90640.MLX90640(i2c)
                        mlx.refresh_rate = refresh_enum
                        print("Thermal camera heropend")
                        status['thermal_connected'] = True
                    except (ValueError, OSError) as e2:
                        print(f"Heropenen mislukt: {e2}")
        except Exception as e:
            print(f"Thermal loop unexpected error: {e}")
            time.sleep(1)

# ============================================
# ROUTES / EVENTS
# ============================================
@app.route('/')
def index():
    return render_template('dashboard.html')

@app.route('/api/status')
def get_status():
    return jsonify(status)
def format_date_be(iso_or_other):
    """
    Converteer een ISO-datum string ('2026-06-08T20:23:19.209Z') naar
    Belgische dag/maand/jaar notatie ('08/06/2026').

    Robuust voor verschillende input-formaten:
      - Volledig ISO met tijd  -> '08/06/2026'
      - ISO zonder Z           -> '08/06/2026'
      - Alleen YYYY-MM-DD      -> '08/06/2026'
      - Lege string / None     -> ''
      - Onparsebaar             -> originele waarde (geen crash)
    """
    if not iso_or_other:
        return ''

    from datetime import datetime

    s = str(iso_or_other).strip()

    # Probeer eerst volledige ISO met tijd (de typische frontend-waarde)
    for fmt in ('%Y-%m-%dT%H:%M:%S.%fZ',
                '%Y-%m-%dT%H:%M:%SZ',
                '%Y-%m-%dT%H:%M:%S',
                '%Y-%m-%d'):
        try:
            dt = datetime.strptime(s, fmt)
            return dt.strftime('%d/%m/%Y')
        except ValueError:
            continue

    # Onparsebaar — return as-is zodat de cel niet leeg is
    return s

@app.route('/api/export-xlsx', methods=['POST'])
def export_xlsx():
    """
    Genereer een gestileerd Excel-bestand uit een lijst log-entries.

    Frontend POST't JSON met:
      {
        "entries": [
          {"lat": ..., "lon": ..., "alt": ..., "time": ..., "date": ...,
           "source": "drone"|"manueel", "status": "...", "notes": "..."},
          ...
        ],
        "filename": "vespatrack_log_20-05-2026_17h30.xlsx"
      }

    Returns het XLSX-bestand als binary download.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    payload = request.get_json(silent=True) or {}
    entries = payload.get('entries', [])
    filename = payload.get('filename', 'vespatrack_log.xlsx')

    # Voor entries zonder adres: probeer alsnog Nominatim aan te roepen.
    # Adressen die binnenkomen worden gecached in coord-log.json zodat
    # volgende export geen Nominatim-calls meer nodig heeft.
    entries, n_filled = ensure_addresses_filled(entries, save_after=True)
    if n_filled > 0:
        print(f"[export] {n_filled} adressen vers opgehaald voor deze export")

    # --- Workbook + sheet aanmaken ---
    wb = Workbook()
    ws = wb.active
    ws.title = "Hornet log"

    # --- Headers definiëren ---
    headers = [
        '#', 'Tijd', 'Datum', 'Bron', 'Status',
        'Adres', 'Hoogte (m)', 'Notitie', 'Google Maps'
    ]

    # --- Header-stijl: vet wit op donkerblauw, gecentreerd ---
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFFFF')
    header_fill = PatternFill(start_color='FF16213E', end_color='FF16213E', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin', color='FFCCCCCC'),
        right=Side(style='thin', color='FFCCCCCC'),
        top=Side(style='thin', color='FFCCCCCC'),
        bottom=Side(style='thin', color='FFCCCCCC')
    )

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    ws.row_dimensions[1].height = 28

    # --- Status-kleuren (afgestemd op dashboard CSS) ---
    status_fills = {
        'gemeld':           PatternFill(start_color='FFE5E7EB', end_color='FFE5E7EB', fill_type='solid'),
        'wordt_onderzocht': PatternFill(start_color='FFFEF3C7', end_color='FFFEF3C7', fill_type='solid'),
        'waargenomen':      PatternFill(start_color='FFD1FAE5', end_color='FFD1FAE5', fill_type='solid'),
        'bestreden':        PatternFill(start_color='FFA7F3D0', end_color='FFA7F3D0', fill_type='solid'),
        'vals_alarm':       PatternFill(start_color='FFE5E7EB', end_color='FFE5E7EB', fill_type='solid'),
    }
    status_labels = {
        'gemeld':           'Gemeld',
        'wordt_onderzocht': 'Wordt onderzocht',
        'waargenomen':      'Waargenomen',
        'bestreden':        'Bestreden',
        'vals_alarm':       'Vals alarm',
        '':                 '',
    }
    source_labels = {
        'drone':   '🚁 drone',
        'manueel': '📍 manueel',
    }

    # --- Data rows ---
    for row_idx, e in enumerate(entries, start=2):
        lat = float(e.get('lat', 0))
        lon = float(e.get('lon', 0))
        alt = float(e.get('alt', 0))
        gmaps_url = f"https://www.google.com/maps?q={lat},{lon}"
        status = e.get('status', '')
        source = e.get('source', '')
        address = e.get('address', '')

        # Fallback bij ontbrekend adres: toon lat/lon zodat bestrijder
        # nog steeds iets concreets ziet. Beter dan een lege cel.
        if not address:
            address = f"({lat:.5f}, {lon:.5f})"

        ws.cell(row=row_idx, column=1, value=row_idx - 1)
        ws.cell(row=row_idx, column=2, value=e.get('time', ''))
        ws.cell(row=row_idx, column=3, value=format_date_be(e.get('date', '')))
        ws.cell(row=row_idx, column=4, value=source_labels.get(source, source))
        ws.cell(row=row_idx, column=5, value=status_labels.get(status, status))
        ws.cell(row=row_idx, column=6, value=address)
        ws.cell(row=row_idx, column=7, value=alt)
        ws.cell(row=row_idx, column=8, value=e.get('notes', ''))

        # Google Maps hyperlink — bewaard zodat bestrijder kan navigeren
        maps_cell = ws.cell(row=row_idx, column=9, value='Open in Maps')
        maps_cell.hyperlink = gmaps_url
        maps_cell.font = Font(color='FF0563C1', underline='single')

        # Status-cel kleur
        if status in status_fills:
            ws.cell(row=row_idx, column=5).fill = status_fills[status]

        # Borders + alignment voor alle cellen in deze rij
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = thin_border
            if col_idx in (1, 2, 3, 4, 5):
                # Tijd/Datum/Bron/Status/# gecentreerd
                cell.alignment = Alignment(horizontal='center', vertical='center')
            elif col_idx == 7:
                # Hoogte gecentreerd
                cell.alignment = Alignment(horizontal='center', vertical='center')
            else:
                # Adres, Notitie, Google Maps links uitgelijnd met wrap
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

        # Hoogte met 2 decimalen
        ws.cell(row=row_idx, column=7).number_format = '0.00'

    # --- Kolombreedtes ---
    column_widths = {
        1: 5,    # #
        2: 11,   # Tijd
        3: 13,   # Datum (DD/MM/YYYY)
        4: 14,   # Bron
        5: 18,   # Status
        6: 40,   # Adres (ruimte voor 'Sint-Pieters-Leeuw, Mechelsgatstraat 12')
        7: 11,   # Hoogte
        8: 35,   # Notitie
        9: 16,   # Google Maps link
    }


    for col_idx, width in column_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # --- Freeze panes (header rij vastzetten bij scroll) ---
    ws.freeze_panes = 'A2'

    # --- Schrijf naar in-memory buffer + stuur als download ---
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

    # ============================================
# REST ENDPOINTS — coördinaat log
# ============================================

@app.route('/api/log', methods=['GET'])
def api_log_get():
    """
    Haal alle gelogde entries op. Wordt door frontend aangeroepen bij
    page-load om de log te tonen.
    """
    entries = load_log()
    return jsonify(entries)


@app.route('/api/log', methods=['POST'])
def api_log_post():
    """
    Voeg een nieuwe entry toe. Verwacht JSON body:
      {lat, lon, alt, time, date, source, status, notes}
    Server genereert het 'id' veld en voegt de entry achteraan toe.

    Bij opslag wordt geprobeerd het adres op te halen via Nominatim
    reverse-geocode (2s timeout). Bij geen internet / timeout blijft
    het address-veld leeg; export-tijd zal het dan alsnog proberen.

    Returns de aangemaakte entry inclusief id en eventueel address.
    """
    payload = request.get_json(silent=True) or {}

    # Minimaal validatie — lat/lon zijn vereist
    if 'lat' not in payload or 'lon' not in payload:
        return jsonify({'error': 'lat en lon zijn vereist'}), 400

    lat = float(payload.get('lat', 0))
    lon = float(payload.get('lon', 0))

    # Probeer adres op te halen — bij falen blijft veld leeg
    address = reverse_geocode(lat, lon)
    if address:
        print(f"[log] adres opgehaald: {address}")

    entry = {
        'id':      generate_entry_id(),
        'lat':     lat,
        'lon':     lon,
        'alt':     float(payload.get('alt', 0)),
        'time':    payload.get('time', ''),
        'date':    payload.get('date', ''),
        'source':  payload.get('source', 'manueel'),
        'status':  payload.get('status', ''),
        'notes':   payload.get('notes', ''),
        'address': address or '',
    }

    entries = load_log()
    entries.append(entry)
    if save_log(entries):
        return jsonify(entry), 201
    return jsonify({'error': 'opslaan mislukt'}), 500


@app.route('/api/log/<entry_id>', methods=['PUT'])
def api_log_put(entry_id):
    """
    Bewerk een bestaande entry. Identificatie via stabiele server-side
    ID, niet via array-index. Body bevat de velden die gewijzigd worden;
    onbenoemde velden blijven ongewijzigd. Returns de bijgewerkte entry.
    """
    payload = request.get_json(silent=True) or {}

    entries = load_log()
    for entry in entries:
        if entry.get('id') == entry_id:
            # Update alleen toegestane velden
            for field in ('status', 'notes', 'lat', 'lon', 'alt'):
                if field in payload:
                    if field in ('lat', 'lon', 'alt'):
                        entry[field] = float(payload[field])
                    else:
                        entry[field] = payload[field]
            if save_log(entries):
                return jsonify(entry)
            return jsonify({'error': 'opslaan mislukt'}), 500

    return jsonify({'error': f'entry {entry_id} niet gevonden'}), 404


@app.route('/api/log/<entry_id>', methods=['DELETE'])
def api_log_delete(entry_id):
    """
    Verwijder één entry op basis van stabiele ID.
    Returns 204 No Content bij succes, 404 als de entry niet bestaat.
    """
    entries = load_log()
    new_entries = [e for e in entries if e.get('id') != entry_id]

    if len(new_entries) == len(entries):
        return jsonify({'error': f'entry {entry_id} niet gevonden'}), 404

    if save_log(new_entries):
        return '', 204
    return jsonify({'error': 'opslaan mislukt'}), 500


@app.route('/api/log', methods=['DELETE'])
def api_log_clear():
    """
    Wis alle entries (equivalent van 'Wissen' knop in dashboard).
    Returns 204 No Content.
    """
    if save_log([]):
        return '', 204
    return jsonify({'error': 'opslaan mislukt'}), 500

    # ============================================
# TILE ROUTE — lokaal-eerst, internet-fallback
# ============================================

@app.route('/tiles/<source>/<int:z>/<int:x>/<int:y>.png')
def serve_tile(source, z, x, y):
    """
    Lever een map-tile aan de browser. Twee stappen:

      1. Probeer uit lokale cache (data/tiles/...)
      2. Bij cache-miss: download van externe server, sla op in cache,
         lever aan browser

    Bij geen internet en geen cache: 404. Frontend Leaflet behandelt
    dit als "tile niet beschikbaar" en toont een grijs vlak.
    """
    if source not in TILE_SOURCES:
        return jsonify({'error': f'onbekende source: {source}'}), 400

    # 1. Cache check
    path = tile_cache_path(source, z, x, y)
    if os.path.exists(path):
        # Detecteer JPEG-magic-bytes (ArcGIS levert JPEG) vs PNG
        with open(path, 'rb') as f:
            magic = f.read(3)
        mime = 'image/jpeg' if magic[:3] == b'\xff\xd8\xff' else 'image/png'
        return send_file(path, mimetype=mime)

    # 2. Cache miss — probeer internet
    data, content_type = fetch_tile_from_internet(source, z, x, y)
    if data is None:
        # Geen internet of tile niet bestaand op server
        return '', 404

    # Sla op voor toekomstig gebruik (best-effort, faal stil als disk vol)
    save_tile_to_cache(source, z, x, y, data)

    # Lever direct aan browser zonder een tweede disk-read
    return data, 200, {'Content-Type': content_type}


@app.route('/api/tiles/stats')
def api_tile_stats():
    """
    Statistieken over de tile cache: aantal tiles per source, totale
    disk-grootte. Wordt door dashboard gebruikt om operator te informeren
    hoeveel offline-kaart-data lokaal beschikbaar is.
    """
    stats = {}
    if not os.path.exists(TILE_CACHE_DIR):
        return jsonify({'sources': {}, 'total_bytes': 0, 'total_count': 0})

    total_bytes = 0
    total_count = 0

    for source in TILE_SOURCES.keys():
        source_dir = os.path.join(TILE_CACHE_DIR, source)
        if not os.path.exists(source_dir):
            stats[source] = {'count': 0, 'bytes': 0}
            continue

        count = 0
        size = 0
        for root, dirs, files in os.walk(source_dir):
            for f in files:
                if f.endswith('.png'):
                    count += 1
                    try:
                        size += os.path.getsize(os.path.join(root, f))
                    except OSError:
                        pass
        stats[source] = {'count': count, 'bytes': size}
        total_bytes += size
        total_count += count

    return jsonify({
        'sources': stats,
        'total_bytes': total_bytes,
        'total_count': total_count,
    })

@app.route('/api/tiles/prefetch', methods=['POST'])
def api_tile_prefetch():
    """
    Start een tile-prefetch in de achtergrond. Body:
      {
        "lat": 50.85, "lon": 4.35,
        "radius_km": 2.0,
        "zoom_min": 13, "zoom_max": 16,
        "sources": ["osm", "sat", "hyb"]
      }
    Returns 202 Accepted met initial status, of 409 Conflict als er al
    een prefetch loopt.
    """
    with prefetch_lock:
        if prefetch_state['running']:
            return jsonify({
                'error': 'prefetch is al bezig',
                'state': dict(prefetch_state)
            }), 409

    payload = request.get_json(silent=True) or {}
    try:
        lat = float(payload['lat'])
        lon = float(payload['lon'])
        radius_km = float(payload['radius_km'])
        zoom_min = int(payload['zoom_min'])
        zoom_max = int(payload['zoom_max'])
        sources = payload.get('sources', ['osm', 'sat', 'hyb'])
    except (KeyError, ValueError, TypeError) as e:
        return jsonify({'error': f'ongeldig payload: {e}'}), 400

    if zoom_min < 0 or zoom_max > 19 or zoom_min > zoom_max:
        return jsonify({'error': 'zoom moet 0-19 zijn, min <= max'}), 400

    if radius_km <= 0 or radius_km > 50:
        return jsonify({'error': 'radius moet 0-50 km zijn'}), 400

    # Start background thread
    with prefetch_lock:
        prefetch_state['running'] = True
        prefetch_state['started_at'] = time.time()
        prefetch_state['total'] = 0
        prefetch_state['done'] = 0
        prefetch_state['success'] = 0
        prefetch_state['fail'] = 0
        prefetch_state['message'] = 'Voorbereiden...'

    thread = threading.Thread(
        target=run_prefetch,
        args=(lat, lon, radius_km, zoom_min, zoom_max, sources),
        daemon=True
    )
    thread.start()

    return jsonify({'state': dict(prefetch_state)}), 202


@app.route('/api/tiles/prefetch/status')
def api_tile_prefetch_status():
    """Huidige status van background prefetch. Frontend pollt deze."""
    with prefetch_lock:
        return jsonify(dict(prefetch_state))


@app.route('/api/tiles', methods=['DELETE'])
def api_tile_delete():
    """
    Wis tile-cache. Query param ?source=osm/sat/hyb om één source te wissen,
    geen param = alles wissen.
    Returns 200 met aantal verwijderde bestanden + bespaarde bytes.
    """
    import shutil
    source = request.args.get('source')

    if source is not None and source not in TILE_SOURCES:
        return jsonify({'error': f'onbekende source: {source}'}), 400

    deleted_count = 0
    deleted_bytes = 0

    if source:
        # Alleen één source
        targets = [os.path.join(TILE_CACHE_DIR, source)]
    else:
        # Alles
        targets = [os.path.join(TILE_CACHE_DIR, s) for s in TILE_SOURCES.keys()]

    for target in targets:
        if not os.path.exists(target):
            continue
        # Tel eerst, dan delete
        for root, dirs, files in os.walk(target):
            for f in files:
                if f.endswith('.png'):
                    try:
                        deleted_bytes += os.path.getsize(os.path.join(root, f))
                        deleted_count += 1
                    except OSError:
                        pass
        try:
            shutil.rmtree(target)
        except OSError as e:
            return jsonify({'error': f'kon {target} niet wissen: {e}'}), 500

    return jsonify({
        'deleted_count': deleted_count,
        'deleted_bytes': deleted_bytes,
    })

@socketio.on('connect')
def handle_connect():
    print('Client verbonden')
    socketio.emit('status_update', status)

@socketio.on('disconnect')
def handle_disconnect():
    print('Client verbroken')

@socketio.on('reset_baseline')
def handle_reset_baseline():
    global baseline_reset_requested
    print('Baseline reset aangevraagd')
    baseline_reset_requested = True

    # ============================================
# THERMAL BASELINE HANDLERS (detectie-modus)
# ============================================

@socketio.on('thermal_baseline_set')
def handle_thermal_baseline_set():
    """
    Operator drukt op 'baseline instellen'. We slaan een snapshot op van het
    laatste frame dat thermal_loop heeft gelezen. Frontend gebruikt die
    snapshot daarna om verschil-rendering te doen (alleen pixels boven baseline).
    """
    global thermal_baseline
    if _last_thermal_frame is None:
        socketio.emit('thermal_baseline_result', {
            'success': False,
            'message': 'Geen thermisch beeld beschikbaar'
        })
        return
    thermal_baseline = list(_last_thermal_frame)
    status['thermal_baseline_set'] = True
    print(f"Thermal baseline ingesteld op {len(thermal_baseline)} pixels")
    socketio.emit('thermal_baseline_result', {
        'success': True,
        'message': 'Baseline ingesteld'
    })


@socketio.on('thermal_baseline_clear')
def handle_thermal_baseline_clear():
    """Wis baseline — frontend gaat terug naar normale rendering."""
    global thermal_baseline
    thermal_baseline = None
    status['thermal_baseline_set'] = False
    print("Thermal baseline gewist")
    socketio.emit('thermal_baseline_result', {
        'success': True,
        'message': 'Baseline gewist'
    })

COMMAND_ACK_TIMEOUT = 5.0   # seconden
def send_command_and_wait_ack(command_id, params, command_name):
    """
    Stuur een MAV_CMD via command_long_send en wacht op de COMMAND_ACK die
    mavlink_loop() in de postbus legt. Roept ZELF NOOIT recv_match aan.
 
    Returns (success: bool, message: str).
    """
    from pymavlink import mavutil
 
    with mav_lock:
        mav = mav_connection
    if mav is None:
        return False, "Pixhawk niet verbonden"
 
    try:
        # 1. Zet de postbus klaar VOOR we sturen, zodat we een ack die
        #    supersnel terugkomt niet missen.
        ack_mailbox.arm(command_id)
 
        # 2. Stuur het commando.
        mav.mav.command_long_send(
            mav.target_system, mav.target_component,
            command_id, 0,
            params[0], params[1], params[2], params[3],
            params[4], params[5], params[6],
        )
 
        # 3. Wacht passief op de ack uit de postbus (geen recv_match).
        result = ack_mailbox.wait(COMMAND_ACK_TIMEOUT)
 
        if result is None:
            return False, f"Geen ACK ontvangen voor {command_name}"
 
        result_codes = {
            mavutil.mavlink.MAV_RESULT_ACCEPTED:  (True,  f"{command_name} geaccepteerd"),
            mavutil.mavlink.MAV_RESULT_DENIED:    (False, f"{command_name} geweigerd door Pixhawk"),
            mavutil.mavlink.MAV_RESULT_UNSUPPORTED: (False, f"{command_name} niet ondersteund"),
            mavutil.mavlink.MAV_RESULT_FAILED:    (False, f"{command_name} mislukt"),
            mavutil.mavlink.MAV_RESULT_TEMPORARILY_REJECTED: (False, f"{command_name} tijdelijk geweigerd (check arming/GPS)"),
        }
        return result_codes.get(result, (False, f"Onbekende ACK: {result}"))
 
    except Exception as e:
        return False, f"MAVLink fout: {e}"


@socketio.on('arm_drone')
def handle_arm_drone():
    """Vraag Pixhawk om te armen. MAV_CMD_COMPONENT_ARM_DISARM met param1=1."""
    from pymavlink import mavutil
    print('ARM commando ontvangen van frontend')
    success, message = send_command_and_wait_ack(
        mavutil.mavlink.MAV_CMD_COMPONENT_ARM_DISARM,
        [1, 0, 0, 0, 0, 0, 0],   # param1=1 = arm
        'ARM'
    )
    socketio.emit('command_result', {'success': success, 'message': message})


@socketio.on('disarm_drone')
def handle_disarm_drone():
    """Vraag Pixhawk om te disarmen. MAV_CMD_COMPONENT_ARM_DISARM met param1=0."""
    from pymavlink import mavutil
    print('DISARM commando ontvangen van frontend')
    success, message = send_command_and_wait_ack(
        mavutil.mavlink.MAV_CMD_COMPONENT_ARM_DISARM,
        [0, 0, 0, 0, 0, 0, 0],   # param1=0 = disarm
        'DISARM'
    )
    socketio.emit('command_result', {'success': success, 'message': message})


def _mode_id_to_name(mode_id):
    """Mode-ID naar naam, spiegelt COPTER_MODES in mavlink_loop."""
    names = {0:'STABILIZE',2:'ALT_HOLD',4:'GUIDED',5:'LOITER',6:'RTL',9:'LAND'}
    return names.get(mode_id, f'MODE_{mode_id}')

def change_mode(mode_id, mode_name=None):
    """
    Wijzig de flight mode en bevestig via status['flight_mode'].
    Roept ZELF NOOIT recv_match aan.
    """
    if mode_name is None:
        mode_name = _mode_id_to_name(mode_id)

    with mav_lock:
        mav = mav_connection
    if mav is None:
        return False, 'Pixhawk niet verbonden'

    try:
        mav.set_mode(mode_id)
        deadline = time.time() + COMMAND_ACK_TIMEOUT
        while time.time() < deadline:
            if status.get('flight_mode') == mode_name:
                return True, f'{mode_name} actief'
            time.sleep(0.1)
        return False, f'{mode_name} niet bevestigd binnen {COMMAND_ACK_TIMEOUT:.0f}s'
    except Exception as e:
        return False, f'MAVLink fout: {e}'


@socketio.on('set_mode')
def handle_set_mode(data):
    """Wijzig flight mode. Frontend stuurt {'mode': <int>}."""
    mode_id = data.get('mode')
    print(f'SET_MODE ontvangen: mode_id={mode_id}')
    success, message = change_mode(mode_id)
    socketio.emit('command_result', {'success': success, 'message': message})


        # ============================================
# MISSIE + NOODKNOPPEN (demo autonome vlucht)
# ============================================

@socketio.on('start_mission')
def handle_start_mission(data=None):
    """
    START MISSIE-knop. Start de zoeksequentie in een aparte thread.

    Frontend stuurt {'altitude': <float>} mee — de zoekhoogte uit het
    dashboard-invoerveld. De hoogte wordt server-side opnieuw geclampt
    in mission.py; de browser-clamp is alleen UX.

    data=None als default zodat een emit zonder payload (oude client,
    of een handmatige emit vanuit de console) niet crasht.
    """
    payload = data or {}
    altitude = payload.get('altitude')
    print(f'START MISSIE ontvangen van frontend (hoogte {altitude} m)')
    success, message = mission.start_mission(
        status, get_mav_connection, socketio.emit, altitude
    )
    socketio.emit('command_result', {'success': success, 'message': message})


@socketio.on('mission_stop_hang')
def handle_stop_hang():
    """STOP & HANG-knop -> LOITER. Drone blijft hangen op huidige positie."""
    print('STOP & HANG (LOITER) ontvangen')
    success, message = change_mode(5, 'LOITER')
    socketio.emit('command_result', {'success': success, 'message': message})
 
@socketio.on('mission_rtl')
def handle_mission_rtl():
    """RTL-knop -> terug naar opstijgpunt en landen."""
    print('RTL ontvangen')
    success, message = change_mode(6, 'RTL')
    socketio.emit('command_result', {'success': success, 'message': message})
 
@socketio.on('mission_land')
def handle_mission_land():
    """LAND NU-knop -> dalen en landen op huidige plek."""
    print('LAND NU ontvangen')
    success, message = change_mode(9, 'LAND')
    socketio.emit('command_result', {'success': success, 'message': message})



@socketio.on('shutdown')
def handle_shutdown():
    print('Shutdown aangevraagd')
    socketio.emit('shutdown_status', {'message': 'Pi wordt afgesloten...'})
    time.sleep(1)
    os.system('sudo shutdown now')

@socketio.on('reboot')
def handle_reboot():
    print('Reboot aangevraagd')
    socketio.emit('shutdown_status', {'message': 'Pi wordt herstart...'})
    time.sleep(1)
    os.system('sudo reboot')


if __name__ == '__main__':
    print("=" * 50)
    print("Hornet Tracker Ground Station v3 (ultimate)")
    print(f"  Source: {SIGNAL_SOURCE}")
    print(f"  Peak-hold: {PEAK_HOLD_SAMPLES} samples (~{PEAK_HOLD_SAMPLES*NUM_SAMPLES/SAMPLE_RATE*1000:.0f} ms window)")
    print(f"  Detect BW: {DETECT_BANDWIDTH/1e3:.0f} kHz")
    print(f"  Threshold: +{THRESHOLD} dB")
    print("=" * 50)

    threading.Thread(target=signal_loop, daemon=True).start()
    threading.Thread(target=wifi_loop, daemon=True).start()
    threading.Thread(target=mavlink_loop, daemon=True).start()
    threading.Thread(target=thermal_loop, daemon=True).start()

    print("Dashboard: http://192.168.1.6:5000")
    print("          http://192.168.4.1:5000")
    print("=" * 50)
    socketio.run(app, host='0.0.0.0', port=5000,
                 debug=False, allow_unsafe_werkzeug=True)